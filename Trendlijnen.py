import matplotlib.pyplot as plt
import pandas as pd 
import numpy as np 
from openpyxl import load_workbook
from scipy import integrate, interpolate

import warnings
warnings.filterwarnings("ignore")


#containers
Cl=6.06   #container lengte
Cb=2.44  #container breedte
Ch=2.59   #container hoogte
Cw=30E3  #container weight

#aantal containers
n=234
#aantal rijen in de hoogte
atiers=3
#aantal containers in breedte
arij=13
#aantal containers in lengte
abay= 6

tp_factor = 66


def beladen(df):
    rho_staal = 7.85E3
    E_staal=210000*(10**6)
    rho_water = 1.025E3
    g = 9.81
    nul = np.zeros(1)
    Loa= df.iloc[0,1] +df.iloc[67,0]
    eind = np.array([Loa])
    onderwater= df.iloc[42:64,0]

    #het gewicht van de containers punt belasting
    G_cont=n*Cw*g

    x=np.arange(0,Loa,0.05)

    #opwaartsekracht verdeelde belasting
    p=np.zeros(len(x))

    p = -df.iloc[42:64,1]*rho_water*g
    p = np.append(p,nul)
    p = np.append(nul,p)
    x_p = df.iloc[42:64,0]
    x_p = np.append(0,x_p)
    x_p = np.append(x_p,max(onderwater))
    p_func = interpolate.interp1d(x_p,p)

    #gewicht verdeelde belasting
    G = df.iloc[101:123,2]*rho_staal*g*tp_factor
    #G=np.append(G,nul)
    x_G = df.iloc[101:123,0]
    #x_G=np.append(x_G,eind)
    G_func = interpolate.interp1d(x_G,G)

    #traagheidsmoment over de lengte
    I = df.iloc[101:123,7]*tp_factor
    #I=np.append(I,nul)
    x_I = df.iloc[101:123,0]
    #x_I=np.append(x_I,eind)
    I_func = interpolate.interp1d(x_I,I)


    G = G_func(x)
    I= I_func(x)
    p= np.zeros(len(x))
    # for loop zodat nadat het onderwater stopt p altijd 0
    for i in range(len(x)):
        if x[i] < min(onderwater):
            p[i] = 0
        elif x[i] > max(onderwater):
            p[i]=0
        else:
            p[i]=p_func(x[i])

    #index=[np.arange(0,1)]

    #I=np.delete(I_func(x),index)

    #Ballast tank
    V_tank=df.iloc[32,1]
    G_tank=V_tank*rho_water*g
    arm_tank= df.iloc[34,1] 

    tank = df.iloc[92:97,1]*rho_water*g#*(df.iloc[35,1]/100)
    x_tank = df.iloc[92:97,0]
    tank_func=interpolate.interp1d(x_tank,tank)

    tanklast=np.zeros(len(x))

    for i in range(len(x)):
        if x[i] > min(x_tank) and x[i] < max(x_tank):
            tanklast[i]= tank_func(x[i])
        else:
            tanklast[i] = 0




    #last op platfrom uitrekenen	

    #som van de krachten
    G_punt=integrate.simpson(G,x)
    P_punt=integrate.simpson(p,x)
    F_c=G_cont
    F_tank=integrate.simpson(tanklast,x)

    F_last = -P_punt  -G_punt  -F_c  -F_tank


    # som van momenten
    x_tank = df.iloc[33,1]
    x_last = 13.5
    COB = df.iloc[20,1]
    COV = df.iloc[21,1]
    #tijdelijke arm
    #arm_c=216
    arm_c = -1*(P_punt*COB +G_punt*COV +F_tank*x_tank +F_last*x_last)/F_c


    #verdeeldebelasting container

    start_cont=arm_c-(0.5*abay*Cl)
    eind_cont=arm_c+(0.5*abay*Cl)

    x_vd=np.arange(start_cont,eind_cont,0.5)
    G_cont_overlengte=G_cont/(eind_cont-start_cont)

    vb_cont=np.zeros(len(x))

    for i in range(len(x)):
        if x[i] > (start_cont) and x[i] < (eind_cont):
            vb_cont[i] = G_cont_overlengte
        else:
            vb_cont[i]= 0 


    #verdeeldebelasting last op platform
    x_platform=[11.8,16.2]
    F_last_overlengte=F_last/(x_platform[1]-x_platform[0])

    vb_last=np.zeros(len(x))

    for i in range(len(x)):
        if x[i] > min(x_platform) and x[i] < max(x_platform):
            vb_last[i] = F_last_overlengte
        else:
            vb_last[i]= 0      



    #verdeelde belasting
    q= p+G+vb_cont+tanklast+vb_last

    # integratie lijnen
    V = integrate.cumtrapz(q,x,initial=0) 
    M = integrate.cumtrapz(V,x,initial=0)

    #Waarde en locatie maximaal moment
    Mmax_index = np.argmax(M)
    M_max = M[Mmax_index] #Waarde maximaal moment
    Loc_M_max = x[Mmax_index] #Locatie maximaal moment

    #Hoekverdraaiing zonder integratie constante
    thetaEI=integrate.cumtrapz(M/(E_staal*I),x,initial=0)

    #Integratie constante
    vEI=integrate.cumtrapz(thetaEI,x,initial=0)
    theta_Mmax = thetaEI[Mmax_index]
    C=(np.max(vEI))/Loa

    #Uiteindelijke verdraaiingslijn
    theta= thetaEI - C

    #Doorbuiging met integratie constante
    v=integrate.cumtrapz(theta,x,initial=0)

    #Waarde en locatie maximale doorbuiging
    vmax_index = np.argmin(v)
    v_max = v[vmax_index] #Waarde maximale doorbuiging
    Loc_v_max = x[vmax_index] #Locatie maximale doorbuiging

    # Maximaal toelaatbaar moment
    sigma_maxtoelaatbaar=190E6
    I_midship=df.iloc[114, 7]*tp_factor
    H=df.iloc[2,1]
    KG_y=df.iloc[21,3]
    y=H-KG_y+(tp_factor*0.001)

    moment_max=(sigma_maxtoelaatbaar*I_midship)/y

    #Weerstandsmoment
    y_boven=df.iloc[101:123,10]-df.iloc[101:123,5]
    y_onder=df.iloc[101:123,5]-df.iloc[101:123,9]
    W=df.iloc[101:123,7]*tp_factor/y_onder

    x_W = df.iloc[101:123,0]
    W_func = interpolate.interp1d(x_W,W)
    W = W_func(x)


    #Spanningsverdeling

    sigma=np.zeros(len(x))

    for i in range(len(x)):
        if M[i]>0:
            sigma[i]=M[i]/W[i]
        else:
            sigma[i] = 0

    sigma_max=np.max(sigma)



    rho_staal = 7.85E3
    E_staal=210E9
    rho_water = 1.025E3
    g = 9.81
    H=df.iloc[2,1]
    nul = np.zeros(1)
    Loa= df.iloc[0,1] +df.iloc[67,0]
    eind = np.array([Loa])
    onderwater= df.iloc[42:64,0]
    Lwl=df.iloc[4,1]

    LCB = df.iloc[20,1]
    dp_leeg=P_punt/(rho_water*g)*-1



    #GM dwarsrichting
    It_x = df.iloc[27,1]
    displacement = df.iloc[18,1]

    KB = df.iloc[20,3]
    KG = df.iloc[21,3]

    #berekening displacement nieuw nadat containers erop zijn
    gewichtschip=displacement*rho_water


    BM_t = It_x/displacement
    KGcont=H+(Ch*atiers/2)
    KGtank=df.iloc[33,3]
    KGlast=H+2.7

    KG_nieuw= (KG*G_punt/g+KGcont*n*Cw+KGlast*F_last/g+V_tank*rho_water*KGtank)/(G_punt/g+n*Cw+F_last/g+V_tank*rho_water)

    #vloeistof reductie
    I_water=df.iloc[38,1]
    gg1=I_water/displacement
    GM_t = KB + BM_t - KG_nieuw-gg1

    #LCG
    LCF = df.iloc[26,1]
    LCGNieuw=(LCF*G_punt+arm_c*n*Cw+x_last*F_last/g+x_tank*V_tank*rho_water)/(G_punt+n*Cw+F_last/g+V_tank*rho_water)

    #GM langsrichting
    It_y = df.iloc[27,2]
    BM_l = It_y/displacement
    GM_l = KB +BM_l-KG_nieuw-gg1

    #momentstelling stabiliteit
    trim_max = 7/180*np.pi   #of negatieve trimhoek
    theta = trim_max/Lwl

    Msl=rho_water*g*displacement*GM_l*(theta)
    #Msl=rho_water*g*displacement*GM_l*theta
    BB1=It_y/displacement*np.tan(theta)

    return [F_last,GM_t,arm_c,max(sigma),df.iloc[150,3]]
    

def trendplotinelkaar(filepath,title):
    datalabel = [r"$Last$", r"$GM_{t}$", r"$LCG_{containers}$",r"$\sigma_{max}$",r"$R_{tot,operationeel}$"]
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    variable = wb.sheetnames
    data = np.zeros(len(datalabel))
    for i in variable:
        df = pd.read_excel(filepath, i)    
        data = np.vstack((data, beladen(df)))

    data = data[1:,]

    variable = [float(numeric_string) for numeric_string in variable]
    figure, ax = plt.subplots(figsize=(10,9))
#    figure.subplots_adjust(right=0.75)


    twin1 = ax.twinx()
    twin2 = ax.twinx()
    twin3 = ax.twinx()
    twin4 = ax.twinx()

    # Offset the right spine of twin2.  The ticks and label have already been
    # placed on the right by twinx above.
    twin1.spines.right.set_position(("axes", 1.0))
    twin2.spines.right.set_position(("axes", 1.1))
    twin3.spines.right.set_position(("axes", 1.2))
    twin4.spines.right.set_position(("axes", 1.3))


    p1, = ax.plot(variable,  data[:,0], label=datalabel[0],c="orange")
    p2, = twin1.plot(variable, data[:,1], label=datalabel[1],c="r")
    p3, = twin2.plot(variable,  data[:,2],label=datalabel[2],c="g")
    p4, = twin3.plot(variable,  data[:,3], label=datalabel[3],c="b")
    p5, = twin4.plot(variable,  data[:,4], label=datalabel[4],c="grey")
    ax.set(ylabel=datalabel[0],ylim=(min(data[:,0]),max(data[:,0])))
    twin1.set(ylabel=datalabel[1],ylim=(min(data[:,1]),max(data[:,1])))
    twin2.set(ylabel=datalabel[2],ylim=(min(data[:,2]),max(data[:,2])))
    twin3.set(ylabel=datalabel[3],ylim=(min(data[:,3]),max(data[:,3])))
    try:
        twin4.set(ylabel=datalabel[4],ylim=(min(data[:,4]),max(data[:,4])))
    except ValueError:
        twin4.set(ylabel=datalabel[4],ylim=(0,6000))

    ax.yaxis.label.set_color(p1.get_color())
    twin1.yaxis.label.set_color(p2.get_color())
    twin2.yaxis.label.set_color(p3.get_color())
    twin3.yaxis.label.set_color(p4.get_color())
    twin4.yaxis.label.set_color(p5.get_color())

    ax.tick_params(axis='y', colors=p1.get_color())
    twin1.tick_params(axis='y', colors=p2.get_color())
    twin2.tick_params(axis='y', colors=p3.get_color())
    twin3.tick_params(axis='y', colors=p4.get_color())
    twin4.tick_params(axis='y', colors=p5.get_color())

#    ax.legend(handles=[p1, p2, p3,p4])
    ax.grid()
    ax.set_title(title)
    plt.tight_layout()
    plt.savefig(r".\variatie onderzoek\ "+title+".png")
'''
trendplotinelkaar(r".\variatie onderzoek\bilgeradius verandering.xlsx","bilge")
trendplotinelkaar(r".\variatie onderzoek\breedte verandering.xlsx","breedte")
trendplotinelkaar(r".\variatie onderzoek\diepgangs verandering.xlsx","diepgang")
trendplotinelkaar(r".\variatie onderzoek\holte verandering.xlsx","holte")
trendplotinelkaar(r".\variatie onderzoek\intrede hoek verandering.xlsx","intrede hoek")
trendplotinelkaar(r".\variatie onderzoek\lengte verandering.xlsx","lengte")
trendplotinelkaar(r".\variatie onderzoek\midship length.xlsx","midscheepse lengte")
trendplotinelkaar(r".\variatie onderzoek\midship verandering.xlsx","midscheepse verandering")
trendplotinelkaar(r".\variatie onderzoek\breedte verandering.xlsx","breedte")
'''

def carene(df):
    rho_staal = 7.85E3
    E_staal=210000*(10**6)
    rho_water = 1.025E3
    g = 9.81
    nul = np.zeros(1)
    Loa= df.iloc[0,1] +df.iloc[67,0]
    eind = np.array([Loa])
    onderwater= df.iloc[42:64,0]

    #het gewicht van de containers punt belasting
    G_cont=n*Cw*g

    x=np.arange(0,Loa,0.05)

    #opwaartsekracht verdeelde belasting
    p=np.zeros(len(x))

    p = -df.iloc[42:64,1]*rho_water*g
    p = np.append(p,nul)
    p = np.append(nul,p)
    x_p = df.iloc[42:64,0]
    x_p = np.append(0,x_p)
    x_p = np.append(x_p,max(onderwater))
    p_func = interpolate.interp1d(x_p,p)

    #gewicht verdeelde belasting
    G = df.iloc[101:123,2]*rho_staal*g*tp_factor
    #G=np.append(G,nul)
    x_G = df.iloc[101:123,0]
    #x_G=np.append(x_G,eind)
    G_func = interpolate.interp1d(x_G,G)

    #traagheidsmoment over de lengte
    I = df.iloc[101:123,7]*tp_factor
    #I=np.append(I,nul)
    x_I = df.iloc[101:123,0]
    #x_I=np.append(x_I,eind)
    I_func = interpolate.interp1d(x_I,I)


    G = G_func(x)
    I= I_func(x)
    p= np.zeros(len(x))
    # for loop zodat nadat het onderwater stopt p altijd 0
    for i in range(len(x)):
        if x[i] < min(onderwater):
            p[i] = 0
        elif x[i] > max(onderwater):
            p[i]=0
        else:
            p[i]=p_func(x[i])

    #index=[np.arange(0,1)]

    #I=np.delete(I_func(x),index)

    #Ballast tank
    V_tank=df.iloc[32,1]
    G_tank=V_tank*rho_water*g
    arm_tank= df.iloc[34,1] 

    tank = df.iloc[92:97,1]*rho_water*g#*(df.iloc[35,1]/100)
    x_tank = df.iloc[92:97,0]
    tank_func=interpolate.interp1d(x_tank,tank)

    tanklast=np.zeros(len(x))

    for i in range(len(x)):
        if x[i] > min(x_tank) and x[i] < max(x_tank):
            tanklast[i]= tank_func(x[i])
        else:
            tanklast[i] = 0




    #last op platfrom uitrekenen	

    #som van de krachten
    G_punt=integrate.simpson(G,x)
    P_punt=integrate.simpson(p,x)
    F_c=G_cont
    F_tank=integrate.simpson(tanklast,x)

    F_last = -P_punt  -G_punt  -F_c  -F_tank


    # som van momenten
    x_tank = df.iloc[33,1]
    x_last = 13.5
    COB = df.iloc[20,1]
    COV = df.iloc[21,1]
    #tijdelijke arm
    #arm_c=216
    arm_c = -1*(P_punt*COB +G_punt*COV +F_tank*x_tank +F_last*x_last)/F_c


    #verdeeldebelasting container

    start_cont=arm_c-(0.5*abay*Cl)
    eind_cont=arm_c+(0.5*abay*Cl)

    x_vd=np.arange(start_cont,eind_cont,0.5)
    G_cont_overlengte=G_cont/(eind_cont-start_cont)

    vb_cont=np.zeros(len(x))

    for i in range(len(x)):
        if x[i] > (start_cont) and x[i] < (eind_cont):
            vb_cont[i] = G_cont_overlengte
        else:
            vb_cont[i]= 0 


    #verdeeldebelasting last op platform
    x_platform=[11.8,16.2]
    F_last_overlengte=F_last/(x_platform[1]-x_platform[0])

    vb_last=np.zeros(len(x))

    for i in range(len(x)):
        if x[i] > min(x_platform) and x[i] < max(x_platform):
            vb_last[i] = F_last_overlengte
        else:
            vb_last[i]= 0      



    #verdeelde belasting
    q= p+G+vb_cont+tanklast+vb_last

    # integratie lijnen
    V = integrate.cumtrapz(q,x,initial=0) 
    M = integrate.cumtrapz(V,x,initial=0)

    #Waarde en locatie maximaal moment
    Mmax_index = np.argmax(M)
    M_max = M[Mmax_index] #Waarde maximaal moment
    Loc_M_max = x[Mmax_index] #Locatie maximaal moment

    #Hoekverdraaiing zonder integratie constante
    thetaEI=integrate.cumtrapz(M/(E_staal*I),x,initial=0)

    #Integratie constante
    vEI=integrate.cumtrapz(thetaEI,x,initial=0)
    theta_Mmax = thetaEI[Mmax_index]
    C=(np.max(vEI))/Loa

    #Uiteindelijke verdraaiingslijn
    theta= thetaEI - C

    #Doorbuiging met integratie constante
    v=integrate.cumtrapz(theta,x,initial=0)

    #Waarde en locatie maximale doorbuiging
    vmax_index = np.argmin(v)
    v_max = v[vmax_index] #Waarde maximale doorbuiging
    Loc_v_max = x[vmax_index] #Locatie maximale doorbuiging

    # Maximaal toelaatbaar moment
    sigma_maxtoelaatbaar=190E6
    I_midship=df.iloc[114, 7]*tp_factor
    H=df.iloc[2,1]
    KG_y=df.iloc[21,3]
    y=H-KG_y+(tp_factor*0.001)

    moment_max=(sigma_maxtoelaatbaar*I_midship)/y

    #Weerstandsmoment
    y_boven=df.iloc[101:123,10]-df.iloc[101:123,5]
    y_onder=df.iloc[101:123,5]-df.iloc[101:123,9]
    W=df.iloc[101:123,7]*tp_factor/y_onder

    x_W = df.iloc[101:123,0]
    W_func = interpolate.interp1d(x_W,W)
    W = W_func(x)


    #Spanningsverdeling

    sigma=np.zeros(len(x))

    for i in range(len(x)):
        if M[i]>0:
            sigma[i]=M[i]/W[i]
        else:
            sigma[i] = 0

    sigma_max=np.max(sigma)



    rho_staal = 7.85E3
    E_staal=210E9
    rho_water = 1.025E3
    g = 9.81
    H=df.iloc[2,1]
    nul = np.zeros(1)
    Loa= df.iloc[0,1] +df.iloc[67,0]
    eind = np.array([Loa])
    onderwater= df.iloc[42:64,0]
    Lwl=df.iloc[4,1]

    LCB = df.iloc[20,1]
    dp_leeg=P_punt/(rho_water*g)*-1



    #GM dwarsrichting
    It_x = df.iloc[27,1]
    displacement = df.iloc[18,1]

    KB = df.iloc[20,3]
    KG = df.iloc[21,3]

    #berekening displacement nieuw nadat containers erop zijn
    gewichtschip=displacement*rho_water


    BM_t = It_x/displacement
    KGcont=H+(Ch*atiers/2)
    KGtank=df.iloc[33,3]
    KGlast=H+2.7

    KG_nieuw= (KG*G_punt/g+KGcont*n*Cw+KGlast*F_last/g+V_tank*rho_water*KGtank)/(G_punt/g+n*Cw+F_last/g+V_tank*rho_water)

    #vloeistof reductie
    I_water=df.iloc[38,1]
    gg1=I_water/displacement
    GM_t = KB + BM_t - KG_nieuw-gg1

    #LCG
    LCF = df.iloc[26,1]
    LCGNieuw=(LCF*G_punt+arm_c*n*Cw+x_last*F_last/g+x_tank*V_tank*rho_water)/(G_punt+n*Cw+F_last/g+V_tank*rho_water)

    #GM langsrichting
    It_y = df.iloc[27,2]
    BM_l = It_y/displacement
    GM_l = KB +BM_l-KG_nieuw-gg1

    #momentstelling stabiliteit
    trim_max = 7/180*np.pi   #of negatieve trimhoek
    theta = trim_max/Lwl

    Msl=rho_water*g*displacement*GM_l*(theta)
    #Msl=rho_water*g*displacement*GM_l*theta
    BB1=It_y/displacement*np.tan(theta)

    return [displacement*10**-4,KB,LCB/10,KG_nieuw/10,LCF/10,It_x*10**-5 ,GM_t/10,It_y*10**-6,GM_l*10**-2]
    

def careneplot(filepath):
    datalabel = [r"$\Delta \; (10^4)$",r"$KB$",r"$LCB\; (10^1)$",r"$KG_{nieuw}\; (10^1)$",r"$LCF\; (10^1)$",r"$I_{t,x} \; (10^5)$",r"$GM_t\; (10^1)$",r"$I_{t,y} \; (10^6)$",r"$GM_l\; (10^2)$"]
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    variable = wb.sheetnames
    data = np.zeros(len(datalabel))
    for i in variable:
        df = pd.read_excel(filepath, i)    
        data = np.vstack((data, carene(df)))
    data = data[1:,]

    variable = [float(numeric_string) for numeric_string in variable]
    
    figure = plt.figure(figsize=(10,10))
    ax = plt.subplot(111)

    for i in range(len(datalabel)):
        plt.plot(data[:,i],variable, label=datalabel[i])
    ax.set_xlabel('[m]')
    ax.set_ylabel('diepgang [m]')
    plt.grid()
    plt.title("Carène diagram")
    plt.legend()
    # Shrink current axis's height by 10% on the bottom
    box = ax.get_position()
    ax.set_position([box.x0, box.y0 + box.height * 0.1,
                    box.width, box.height * 0.9])
    # Put a legend below current axis
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1),
            fancybox=True, shadow=True, ncol=3)

    plt.savefig(r".\variatie onderzoek\Carene.png")


careneplot(r".\variatie onderzoek\Carene the1 v1.2.xlsx")
